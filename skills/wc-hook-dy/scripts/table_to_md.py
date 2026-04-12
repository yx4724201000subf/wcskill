#!/usr/bin/env python3
"""
table_to_md.py — 把常见表格文件（csv / tsv / xlsx）转成 Markdown 表格。

用法：
  python3 table_to_md.py <input_file> [--sheet SHEET_NAME] [--out OUT.md] [--max-cell N]

支持格式：.csv .tsv .xlsx .xlsm
"""
import sys
import csv
import argparse
from pathlib import Path


def csv_to_rows(path, delimiter=","):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [list(r) for r in csv.reader(f, delimiter=delimiter)]


def xlsx_to_rows(path, sheet_name=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("错误：读取 xlsx 需要 openpyxl。安装：pip install openpyxl\n")
        sys.exit(2)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])
    return rows


def rows_to_md(rows, max_cell=None):
    def clean(x):
        if x is None:
            return ""
        s = str(x).replace("\r", " ").replace("\n", " ").replace("|", "\\|").strip()
        if max_cell and len(s) > max_cell:
            s = s[: max_cell - 1] + "…"
        return s

    rows = [[clean(c) for c in r] for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return ""
    max_cols = max(len(r) for r in rows)
    for r in rows:
        while len(r) < max_cols:
            r.append("")
    header, body = rows[0], rows[1:]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * max_cols) + " |",
    ]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("input")
    p.add_argument("--sheet", default=None, help="xlsx sheet 名；默认取 active sheet")
    p.add_argument("--out", default=None, help="输出文件路径；不填则打印到 stdout")
    p.add_argument(
        "--max-cell",
        type=int,
        default=None,
        help="单格最长字符数（超出截断，仅预览用）；不填保留原文",
    )
    args = p.parse_args()
    path = Path(args.input)
    if not path.exists():
        sys.stderr.write(f"找不到文件: {path}\n")
        sys.exit(1)
    ext = path.suffix.lower()
    if ext == ".csv":
        rows = csv_to_rows(path, ",")
    elif ext == ".tsv":
        rows = csv_to_rows(path, "\t")
    elif ext in (".xlsx", ".xlsm"):
        rows = xlsx_to_rows(path, args.sheet)
    else:
        sys.stderr.write(f"不支持的文件格式: {ext}（支持 .csv .tsv .xlsx .xlsm）\n")
        sys.exit(1)
    md = rows_to_md(rows, args.max_cell)
    if args.out:
        Path(args.out).write_text(md, encoding="utf-8")
        print(f"已写出: {args.out}（{len(rows)} 行 / {len(rows[0]) if rows else 0} 列）")
    else:
        print(md)


if __name__ == "__main__":
    main()
